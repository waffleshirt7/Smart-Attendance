#!/usr/bin/env python3
"""
Semester-wide Attendance Register
Shows: Roll No. | Name | Day1 | Day2 | Day3 | ... with P/A marks
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_all_attendance():
    """Load all attendance records from JSON files."""
    records_dir = Path("attendance_records")
    
    if not records_dir.exists():
        print("Error: No attendance records found.")
        return None, None
    
    all_records = []
    
    # Load all JSON files
    for json_file in sorted(records_dir.glob("*.json")):
        try:
            with open(json_file, 'r') as f:
                data = json.load(f)
                all_records.extend(data)
        except Exception:
            pass
    
    if not all_records:
        print("Error: No attendance data in records.")
        return None, None
    
    # Create DataFrame
    df = pd.DataFrame(all_records)
    
    # Ensure Date column is present
    if 'Date' not in df.columns:
        print("Error: Date column not found in records.")
        return None, None
    
    # Convert date format (DD-MM-YYYY to sortable format), handle epoch values
    def parse_date(val):
        try:
            # If value is int or looks like epoch
            if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
                # Try ms-since-epoch (pandas sometimes stores as ms)
                v = int(val)
                # Heuristic: if value is very large, treat as ms
                if v > 1e12:
                    return pd.to_datetime(v, unit='ms')
                else:
                    return pd.to_datetime(v, unit='s')
            # Try parsing as string date
            return pd.to_datetime(val, format='%d-%m-%Y', errors='coerce')
        except Exception:
            return pd.NaT

    df['Date'] = df['Date'].apply(parse_date)
    df = df.dropna(subset=['Date'])
    df = df.sort_values('Date')
    return df, sorted(df['Date'].unique())


def get_student_data(df):
    """Extract unique students with their roll numbers."""
    students = df[['Name', 'Roll No.']].drop_duplicates().reset_index(drop=True)
    students = students.sort_values('Roll No.', key=lambda x: pd.to_numeric(x, errors='coerce')).reset_index(drop=True)
    return students


def create_semester_register(df, dates, students):
    """Create semester register with P/A marks."""
    
    # Initialize register with Roll No. and Name
    register = students.copy()
    
    # Add attendance for each date
    for date in dates:
        date_str = date.strftime('%d-%m-%Y')
        date_col = date.strftime('%d/%m')  # Short format for header
        
        attendance = []
        for _, student in students.iterrows():
            name = student['Name']
            roll_no = student['Roll No.']
            # Check if student has attendance on this date
            record = df[(df['Roll No.'] == roll_no) & (df['Date'] == date)]
            if len(record) > 0:
                attendance.append('P')  # Present
            else:
                attendance.append('A')  # Absent
        
        register[date_col] = attendance
    
    # Rename columns
    register = register.rename(columns={'Roll_No': 'Roll No.', 'Name': 'Name'})
    
    return register


def save_to_excel(register, dates):
    # Swap Column A and B (Name ↔ Roll No.)
    cols = register.columns.tolist()
    cols[0], cols[1] = cols[1], cols[0]
    register = register[cols]
    """Save register to Excel with formatting."""
    filename = "attendance_sheets/semester_register.xlsx"
    
    # Create directory if not exists
    Path("attendance_sheets").mkdir(exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Semester Register"
    
    # Write header
    ws['A1'] = "ATTENDANCE REGISTER"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells(f'A1:{get_column_letter(len(register.columns))}1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 25
    
    # Write date generated
    ws['A2'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    
    # Write column headers
    header_row = 3
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, col_name in enumerate(register.columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[header_row].height = 30
    
    # Write data
    for row_idx, row in register.iterrows():
        excel_row = row_idx + header_row + 1
        
        for col_idx, value in enumerate(row.values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Color code: P = Green, A = Red
            if value == 'P':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(bold=True, color="006100")
            elif value == 'A':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True, color="9C0006")
            else:
                if col_idx <= 2:  # Roll No. and Name columns
                    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Roll No.
    ws.column_dimensions['B'].width = 20  # Name
    
    for col_idx in range(3, len(register.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8
    
    # Add legend
    legend_row = len(register) + 5
    ws[f'A{legend_row}'] = "Legend:"
    ws[f'A{legend_row}'].font = Font(bold=True)
    
    ws[f'A{legend_row + 1}'] = "P"
    ws[f'A{legend_row + 1}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws[f'B{legend_row + 1}'] = "Present"
    
    ws[f'A{legend_row + 2}'] = "A"
    ws[f'A{legend_row + 2}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws[f'B{legend_row + 2}'] = "Absent"
    
    wb.save(filename)
    
    return filename


def save_to_csv(register):
    """Save register to CSV."""
    filename = "attendance_sheets/semester_register.csv"
    Path("attendance_sheets").mkdir(exist_ok=True)
    
    register.to_csv(filename, index=False)
    
    return filename


def save_to_html(register, dates):
    """Save register to HTML with styling."""
    filename = "attendance_sheets/semester_register.html"
    Path("attendance_sheets").mkdir(exist_ok=True)
    
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Semester Attendance Register</title>
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 20px;
                min-height: 100vh;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background: white;
                border-radius: 10px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.3);
                padding: 30px;
            }
            h1 {
                text-align: center;
                color: #1F4E78;
                margin-bottom: 10px;
                font-size: 28px;
            }
            .info {
                text-align: center;
                color: #666;
                margin-bottom: 20px;
                font-size: 14px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }
            th {
                background: linear-gradient(135deg, #1F4E78 0%, #2E5090 100%);
                color: white;
                padding: 12px;
                text-align: center;
                font-weight: bold;
                border: 1px solid #ddd;
                font-size: 13px;
            }
            td {
                padding: 10px;
                text-align: center;
                border: 1px solid #ddd;
                font-size: 14px;
            }
            tr:nth-child(even) {
                background-color: #f9f9f9;
            }
            tr:hover {
                background-color: #f0f0f0;
            }
            .roll-no {
                font-weight: bold;
                text-align: center;
                width: 60px;
            }
            .name {
                text-align: left;
                min-width: 200px;
            }
            .present {
                background-color: #C6EFCE;
                color: #006100;
                font-weight: bold;
            }
            .absent {
                background-color: #FFC7CE;
                color: #9C0006;
                font-weight: bold;
            }
            .legend {
                margin-top: 30px;
                padding: 20px;
                background-color: #f5f5f5;
                border-radius: 5px;
            }
            .legend h3 {
                color: #1F4E78;
                margin-bottom: 10px;
            }
            .legend-item {
                display: inline-block;
                margin-right: 30px;
                font-size: 14px;
            }
            .legend-box {
                display: inline-block;
                width: 20px;
                height: 20px;
                margin-right: 8px;
                vertical-align: middle;
                border-radius: 3px;
            }
            .present-box {
                background-color: #C6EFCE;
            }
            .absent-box {
                background-color: #FFC7CE;
            }
            .stats {
                margin-top: 20px;
                padding: 15px;
                background-color: #E8F4F8;
                border-left: 4px solid #2196F3;
                border-radius: 3px;
            }
            .stats p {
                margin: 5px 0;
                color: #333;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📚 ATTENDANCE REGISTER</h1>
            <div class="info">Generated on: """ + datetime.now().strftime('%d-%m-%Y %H:%M:%S') + """</div>
            
            <table>
                <thead>
                    <tr>
                        <th style="width: 80px;">Roll No.</th>
                        <th style="min-width: 200px;">Name</th>
    """
    
    # Add date headers
    for date in dates:
        date_str = date.strftime('%d/%m')
        html_content += f'<th style="width: 50px;">{date_str}</th>'
    
    html_content += """
                    </tr>
                </thead>
                <tbody>
    """
    
    # Add rows
    for _, row in register.iterrows():
        html_content += '<tr>'
        html_content += f'<td class="roll-no">{row["Roll No."]}</td>'
        html_content += f'<td class="name">{row["Name"]}</td>'
        
        # Add P/A marks
        for col in register.columns[2:]:
            value = row[col]
            if value == 'P':
                html_content += f'<td class="present">P</td>'
            else:
                html_content += f'<td class="absent">A</td>'
        
        html_content += '</tr>'
    
    html_content += """
                </tbody>
            </table>
            
            <div class="legend">
                <h3>Legend</h3>
                <div class="legend-item">
                    <span class="legend-box present-box"></span>
                    <span>P = Present</span>
                </div>
                <div class="legend-item">
                    <span class="legend-box absent-box"></span>
                    <span>A = Absent</span>
                </div>
            </div>
            
            <div class="stats">
                <p><strong>Total Students:</strong> """ + str(len(register)) + """</p>
                <p><strong>Total Days:</strong> """ + str(len(register.columns) - 2) + """</p>
                <p><strong>Report Generated:</strong> """ + datetime.now().strftime('%d-%m-%Y %H:%M:%S') + """</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    return filename


def main():
    df, dates = load_all_attendance()
    if df is None:
        return
    students = get_student_data(df)
    register = create_semester_register(df, dates, students)
    save_to_excel(register, dates)
    save_to_csv(register)
    save_to_html(register, dates)
    print(f"Register saved: attendance_sheets/ ({len(register)} students, {len(dates)} days)")


if __name__ == "__main__":
    main()
